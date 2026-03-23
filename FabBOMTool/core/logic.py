"""
core/logic.py
All business logic preserved exactly from the original app.py.
Zero Tkinter dependencies. Pure Python.
"""

import re
import json
import csv
from pathlib import Path

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.security import check_password_hash, generate_password_hash

# Embedded legend (optional - place legend_embedded.py next to this file or in project root)
try:
    import legend_embedded  # type: ignore
except Exception:
    legend_embedded = None

APP_NAME    = "Fabrication BOM Tool"
APP_VERSION = "2026.3.3"
DEFAULT_ADMIN_PASSWORD = "FBT2026!"

# ============================
# Paths
# ============================

def get_app_dir() -> Path:
    return Path(__file__).resolve().parent.parent

APP_DIR                      = get_app_dir()
DATA_DIR                     = APP_DIR / "data"
LEGEND_CACHE_PATH            = DATA_DIR / "Legend.cache.json"
LEGEND_XLSX_PATH             = APP_DIR / "Legend.xlsx"
SETTINGS_FILE                = DATA_DIR / "settings.json"
SETTINGS_BACKUP_FILE         = DATA_DIR / "settings.backup.json"
COMPANY_DEFAULTS_LOCAL_PATH  = DATA_DIR / "CompanyDefaults.json"
EXPORTS_DIR                  = APP_DIR / "exports"

# ============================
# Material Types preset
# ============================

MATERIAL_TYPE_PRESET_RAW = [
    "PVDF","Cast Iron","Forged Steel","Carbon Steel","Malleable Iron","Copper",
    "Bronze","PVC","Stainless Steel","Brass","Cast Steel","Ductile Iron",
    "Polypropylene","PVC SDR35","Polyethylene","PEX","Victaulic (OEM) Imperial",
    "304 Stainless Steel","Cast Brass","Aluminum","Nickel iron","CPVC",
    "Rubber","FPM","Copper B280","Epoxy Steel/Polyethylene","Forged Brass",
    "Steel","Neoprene","Engineered Polymer","Cast Copper","Cast Bronze",
    "Polypropelene","ABS","MDPE","HDPE","Forged Bronze","Black Iron",
    "Forged Carbon Steel","Cast Carbon Steel",
]


def dedupe_case_insensitive_keep_first(values: list) -> list:
    seen = set()
    out = []
    for v in values:
        s = (v or "").strip()
        if not s:
            continue
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


MATERIAL_TYPE_PRESET = dedupe_case_insensitive_keep_first(MATERIAL_TYPE_PRESET_RAW)

# ============================
# Fitting Types
# ============================

_BASE_FITTING_TYPES = [
    "Adapter","Cap","Coupling","Elbow","Flange","Nipple","Olet","Pipe",
    "Reducer","Sleeve","Tee","Union","Valve","Wye","Unclassified",
]


def _legend_bucket_values_from_embedded() -> set:
    out = set()
    try:
        if legend_embedded is None:
            return out
        payload = getattr(legend_embedded, "LEGEND_PAYLOAD", None)
        if not isinstance(payload, dict):
            return out
        for key in ("concat","concat_nospace","desc","desc_nospace"):
            d = payload.get(key, {})
            if isinstance(d, dict):
                for v in d.values():
                    s = str(v or "").strip()
                    if s and s.lower() != "n/a":
                        out.add(s)
    except Exception:
        return set()
    return out


def _build_fitting_types() -> list:
    buckets   = _legend_bucket_values_from_embedded()
    all_types = dedupe_case_insensitive_keep_first([*_BASE_FITTING_TYPES, *sorted(list(buckets))])
    sorted_types = sorted([t for t in all_types if t.lower() != "unclassified"], key=lambda x: x.lower())
    if any(t.lower() == "unclassified" for t in all_types):
        sorted_types.append("Unclassified")
    return sorted_types


FITTING_TYPES = _build_fitting_types()

# ============================
# Regex patterns
# ============================

LENGTH_RE = re.compile(
    r"("
    r"\d+'\s*-\s*[\d\s/]+\""
    r"|"
    r"\d+\s+\d+\s*/\s*\d+\""
    r"|"
    r"\d+\s*-\s*\d+\s*/\s*\d+\""
    r"|"
    r"\d+\s*/\s*\d+\""
    r"|"
    r"\d+(?:\.\d+)?\""
    r")"
)

BATCH_IN_FILENAME_RE   = re.compile(r"\bB\d{3,6}\b", re.IGNORECASE)
BATCH_IN_ROW_RE        = re.compile(r"\bB\d{3,6}\b", re.IGNORECASE)
FALLBACK_BATCH_TOKEN_RE = re.compile(r"[A-Z0-9]{6,}")

PROGRAM_DEFAULT_MULTIPLIER = 2.00
FALLBACK_MULTIPLIER        = 2.00

# ============================
# Hardcoded company multipliers
# ============================

HARDCODED_COMPANY_SIDE_MULTIPLIERS = {
    'PVDF':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Cast Iron':{'Adapter':1.0,'Cap':1.0,'Coupling':0.0,'Elbow':1.0,'Flange':1.0,'Nipple':2.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Forged Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Carbon Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Malleable Iron':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Copper':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Bronze':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'PVC':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Stainless Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Brass':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Cast Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Ductile Iron':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Polypropylene':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'PVC SDR35':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Polyethylene':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'PEX':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Victaulic (OEM) Imperial':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    '304 Stainless Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Cast Brass':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Aluminum':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Nickel iron':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'CPVC':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Rubber':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'FPM':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Copper B280':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Epoxy Steel/Polyethylene':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Forged Brass':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Neoprene':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Engineered Polymer':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Cast Copper':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Cast Bronze':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'ABS':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'MDPE':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'HDPE':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Forged Bronze':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Black Iron':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Forged Carbon Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
    'Cast Carbon Steel':{'Adapter':1.0,'Cap':1.0,'Coupling':1.0,'Elbow':1.0,'Flange':1.0,'Nipple':1.0,'Olet':3.0,'Pipe':0.0,'Reducer':2.0,'Sleeve':0.0,'Tee':2.0,'Union':2.0,'Valve':2.0,'Wye':2.0,'Unclassified':0.0},
}

# ============================
# Helpers
# ============================

def clean_token(x) -> str:
    return str(x).replace("\u201c",'"').replace("\u201d",'"').replace("\n"," ").strip()

def _norm_text(s: str) -> str:
    s = str(s or "").upper()
    s = re.sub(r"[^A-Z0-9]+"," ",s)
    return re.sub(r"\s+"," ",s).strip()

def _norm_text_nospace(s: str) -> str:
    s = str(s or "").upper()
    return re.sub(r"[^A-Z0-9]+","",s).strip()

def _ci_get(d: dict, key: str):
    if not isinstance(d,dict): return None
    k = str(key or "").strip().lower()
    if not k: return None
    for kk,vv in d.items():
        if str(kk).strip().lower() == k:
            return vv
    return None

# ============================
# Legend
# ============================

def build_legend_maps_from_xlsx(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            headers[str(v).strip().lower()] = c
    col_desc   = headers.get("description", 4)
    col_concat = headers.get("description.1", 5)
    col_fit    = headers.get("multiplier reference", 9)
    col_mfg    = headers.get("manufacturer", 3)
    concat_map = {}
    desc_to_fit = {}
    manufacturers = set()
    for r in range(2, ws.max_row + 1):
        desc   = ws.cell(r, col_desc).value
        concat = ws.cell(r, col_concat).value
        fit    = ws.cell(r, col_fit).value
        mfg    = ws.cell(r, col_mfg).value if col_mfg else None
        fit_s  = str(fit or "").strip()
        if not fit_s or fit_s.lower() == "n/a":
            continue
        if mfg:
            manufacturers.add(_norm_text(mfg))
        if concat:
            k = _norm_text(concat)
            if k: concat_map[k] = fit_s
        if desc:
            k2 = _norm_text(desc)
            if k2: desc_to_fit.setdefault(k2, set()).add(fit_s)
    desc_map = {k2: next(iter(fits)) for k2, fits in desc_to_fit.items() if len(fits) == 1}
    return concat_map, desc_map, manufacturers


def load_legend_maps():
    # 1) Embedded
    try:
        if legend_embedded is not None:
            p = getattr(legend_embedded, "LEGEND_PAYLOAD", None)
            if isinstance(p, dict):
                cm  = p.get("concat", {}) or {}
                cmn = p.get("concat_nospace", {}) or {}
                dm  = p.get("desc", {}) or {}
                dmn = p.get("desc_nospace", {}) or {}
                mfg = set((p.get("manufacturers") or p.get("mfg") or []) or [])
                if all(isinstance(x, dict) for x in (cm, cmn, dm, dmn)) and (cm or cmn or dm or dmn):
                    return cm, cmn, dm, dmn, set(map(_norm_text, mfg))
    except Exception:
        pass

    # 2) Cached JSON
    try:
        if LEGEND_CACHE_PATH.exists():
            p = json.loads(LEGEND_CACHE_PATH.read_text(encoding="utf-8"))
            cm  = p.get("concat", {}) or {}
            cmn = p.get("concat_nospace", {}) or {}
            dm  = p.get("desc", {}) or {}
            dmn = p.get("desc_nospace", {}) or {}
            mfg = set((p.get("manufacturers") or p.get("mfg") or []) or [])
            if all(isinstance(x, dict) for x in (cm, cmn, dm, dmn)) and (cm or cmn or dm or dmn):
                return cm, cmn, dm, dmn, set(map(_norm_text, mfg))
    except Exception:
        pass

    # 3) Legacy XLSX fallback
    try:
        if LEGEND_XLSX_PATH.exists():
            cm, dm, mfg = build_legend_maps_from_xlsx(LEGEND_XLSX_PATH)
            cmn = {_norm_text_nospace(k): v for k, v in cm.items() if _norm_text_nospace(k)}
            dmn = {_norm_text_nospace(k): v for k, v in dm.items() if _norm_text_nospace(k)}
            return cm, cmn, dm, dmn, mfg
    except Exception:
        pass

    return {}, {}, {}, {}, set()


def strip_manufacturer_prefix(desc: str, manufacturers_norm: set) -> str:
    raw = str(desc or "").strip()
    if not raw:
        return ""
    n = _norm_text(raw)
    if not n:
        return ""
    parts = n.split(" ")
    for k in range(4, 0, -1):
        prefix = " ".join(parts[:k]).strip()
        if prefix in manufacturers_norm:
            remainder = " ".join(parts[k:]).strip()
            return remainder or n
    return n


def fitting_type_override(description: str) -> str | None:
    u = str(description or "").upper()
    if not u:
        return None
    if "SHORT SWEEP" in u:
        return "Elbow"
    if "COMBINATION" in u:
        return "Tee"
    if "SANITARY CROSS" in u:
        return "Tee"
    if re.search(r"\bP[\s-]?TRAP\b", u):
        return "Tee"
    if "CHARLOTTE NONH 52 S TAPPED FERRULE WITH SOUTHERN RAISED HEAD BRASS PLUG" in _norm_text(description):
        return "Cap"
    return None


def classify_fitting_type(description: str) -> str:
    override = fitting_type_override(description)
    if override:
        return override
    u = str(description or "").upper()
    if "UNION" in u:
        return "Union"
    if "OLET" in u or "WELDOLET" in u or "THREDOLET" in u or "SOCKOLET" in u:
        return "Olet"
    if "FLANGE" in u:
        return "Flange"
    if "SLEEVE" in u:
        return "Sleeve"
    if "NIPPLE" in u:
        return "Nipple"
    if "CAP" in u:
        return "Cap"
    if "PIPE" in u:
        return "Pipe"
    if "COUPLING" in u:
        return "Coupling"
    if re.search(r"\bTEE\b", u):
        return "Tee"
    if "WYE" in u:
        return "Wye"
    if "ELBOW" in u or "ELL" in u or "BEND" in u:
        return "Elbow"
    if "REDUCER" in u or "BUSHING" in u:
        return "Reducer"
    if "VALVE" in u:
        return "Valve"
    if "ADAPTER" in u:
        return "Adapter"
    return "Unclassified"


def classify_fitting_type_with_legend(description: str, legend_maps=None) -> str:
    override = fitting_type_override(description)
    if override:
        return override
    if legend_maps is None:
        legend_maps = load_legend_maps()
    concat_map, concat_map_ns, desc_map, desc_map_ns, manufacturers_norm = legend_maps
    raw = str(description or "").strip()
    if not raw:
        return "Unclassified"
    n1 = _norm_text(raw)
    if n1:
        ft = concat_map.get(n1)
        if ft:
            return ft
    n1n = _norm_text_nospace(raw)
    if n1n:
        ft = concat_map_ns.get(n1n)
        if ft:
            return ft
    stripped = strip_manufacturer_prefix(raw, manufacturers_norm)
    n2 = _norm_text(stripped)
    n2n = _norm_text_nospace(stripped)
    if n2:
        ft = concat_map.get(n2) or desc_map.get(n2)
        if ft:
            return ft
    if n2n:
        ft = concat_map_ns.get(n2n) or desc_map_ns.get(n2n)
        if ft:
            return ft
    return classify_fitting_type(raw)


def classify_fitting(description: str, manufacturers: set, concat_map: dict, concat_map_ns: dict, desc_map: dict, desc_map_ns: dict) -> str:
    return classify_fitting_type_with_legend(
        description,
        (concat_map, concat_map_ns, desc_map, desc_map_ns, manufacturers),
    )

# ============================
# Material normalization / matching
# ============================

def normalize_material_token(s: str) -> str:
    s = str(s or "").upper().replace("&", " AND ")
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.replace("POLYPROPELENE", "POLYPROPYLENE")


def build_material_alias_map(material_types: list) -> dict:
    alias = {}
    if not isinstance(material_types, list):
        return alias
    canon_norm = {}
    for canonical in material_types:
        c = str(canonical or "").strip()
        if not c:
            continue
        n = normalize_material_token(c)
        ns = _norm_text_nospace(c)
        if n:
            canon_norm[n] = c
        alias[n] = c
        alias[ns] = c
        first = n.split(" ")[0] if n else ""
        if first:
            alias.setdefault(first, c)

    def add(a, canonical):
        if not a or not canonical:
            return
        alias[normalize_material_token(a)] = canonical
        alias[_norm_text_nospace(a)] = canonical

    if "COPPER" in canon_norm:
        add("COOPER", canon_norm["COPPER"])
        add("CU", canon_norm["COPPER"])
    if "PVC" in canon_norm:
        add("POLYVINYL CHLORIDE", canon_norm["PVC"])
    for k, v in canon_norm.items():
        if k in ("STAINLESS STEEL", "304 STAINLESS STEEL"):
            add("SS", v)
            add("S S", v)
    if "CAST IRON" in canon_norm:
        add("CI", canon_norm["CAST IRON"])
    if "CARBON STEEL" in canon_norm:
        add("CS", canon_norm["CARBON STEEL"])
    if "COPPER B280" in canon_norm:
        add("COPPER R280", canon_norm["COPPER B280"])
        add("B280", canon_norm["COPPER B280"])
        add("R280", canon_norm["COPPER B280"])
    if "NICKEL IRON" in canon_norm:
        add("NICKELIRON", canon_norm["NICKEL IRON"])
    return alias


def material_type_from_material(material: str, known_material_types: list) -> str:
    raw  = clean_token(material)
    if not raw: return "UNKNOWN"
    mats = [str(x).strip() for x in (known_material_types or []) if str(x).strip()]
    if not mats: return "UNKNOWN"
    norm  = normalize_material_token(raw)
    ns    = _norm_text_nospace(raw)
    alias = build_material_alias_map(mats)
    if norm in alias: return alias[norm]
    if ns   in alias: return alias[ns]
    parts = norm.split()
    for n in range(min(4,len(parts)),0,-1):
        key    = " ".join(parts[:n])
        key_ns = _norm_text_nospace(key)
        if key    in alias: return alias[key]
        if key_ns in alias: return alias[key_ns]
    for m in mats:
        mk = normalize_material_token(m)
        if mk and mk in norm: return m
    first = raw.split()[0].strip()
    return first if first else "UNKNOWN"

# ============================
# Settings / Defaults
# ============================



def sort_case_insensitive(values: list[str]) -> list[str]:
    return sorted(values, key=lambda value: (str(value).casefold(), str(value)))

DEFAULT_SETTINGS = {
    "admin_password_hash":        "",
    "multiplier_mode":            "Company",
    "selected_project":           "",
    "material_types":             MATERIAL_TYPE_PRESET,
    "company_side_multipliers":   {},
    "project_side_multipliers":   {},
    "exclude_fitting_types":      {ft: False for ft in _BASE_FITTING_TYPES},
}


def _clean_multiplier_table(tbl) -> dict:
    if not isinstance(tbl, dict): return {}
    out = {}
    for mat, mp in tbl.items():
        mat_s = str(mat or "").strip()
        if not mat_s or not isinstance(mp, dict): continue
        out[mat_s] = {}
        for ft, val in mp.items():
            ft_s = str(ft or "").strip()
            if not ft_s: continue
            try: out[mat_s][ft_s] = round(float(val), 2)
            except Exception: pass
    return out


def _deep_merge_defaults(data: dict, defaults: dict) -> dict:
    out = dict(defaults)
    for k, v in data.items():
        if k in out and isinstance(out[k], dict) and isinstance(v, dict):
            out[k] = _deep_merge_defaults(v, out[k])
        else:
            out[k] = v
    return out


def build_full_default_table(material_types: list) -> dict:
    out = {}
    for mat in material_types:
        out[mat] = {}
        base_mat = _ci_get(HARDCODED_COMPANY_SIDE_MULTIPLIERS, mat)
        for ft in FITTING_TYPES:
            if isinstance(base_mat, dict) and ft in base_mat:
                out[mat][ft] = round(float(base_mat[ft]), 2)
            else:
                out[mat][ft] = round(float(PROGRAM_DEFAULT_MULTIPLIER), 2)
    return out


def ensure_admin_password_hash(settings: dict) -> dict:
    password_hash = str(settings.get("admin_password_hash", "") or "").strip()
    legacy_password = str(settings.get("admin_password", "") or "").strip()
    if not password_hash:
        seed_password = legacy_password or DEFAULT_ADMIN_PASSWORD
        settings["admin_password_hash"] = generate_password_hash(seed_password)
    settings.pop("admin_password", None)
    return settings


def password_matches(settings: dict, password: str) -> bool:
    ensure_admin_password_hash(settings)
    provided = str(password or "")
    password_hash = str(settings.get("admin_password_hash", "") or "")
    if not password_hash:
        return False
    try:
        return check_password_hash(password_hash, provided)
    except ValueError:
        return False


def set_admin_password(settings: dict, password: str) -> dict:
    settings["admin_password_hash"] = generate_password_hash(str(password or ""))
    settings.pop("admin_password", None)
    return settings


def ensure_company_defaults(settings: dict) -> dict:
    settings = ensure_admin_password_hash(settings)
    mats = settings.get("material_types", [])
    if not isinstance(mats, list) or not mats:
        mats = MATERIAL_TYPE_PRESET
    mats = dedupe_case_insensitive_keep_first([str(x).strip() for x in mats if str(x).strip()])
    mats = sort_case_insensitive(mats)
    settings["material_types"] = mats

    existing = _clean_multiplier_table(settings.get("company_side_multipliers", {}))
    full     = build_full_default_table(mats)
    for mat in mats:
        full.setdefault(mat, {})
        mat_tbl = _ci_get(existing, mat)
        if isinstance(mat_tbl, dict):
            for ft, val in mat_tbl.items():
                try: full[mat][ft] = round(float(val), 2)
                except Exception: pass
        for ft in FITTING_TYPES:
            full[mat].setdefault(ft, round(float(PROGRAM_DEFAULT_MULTIPLIER), 2))
    settings["company_side_multipliers"] = full

    ex = settings.get("exclude_fitting_types", {})
    if not isinstance(ex, dict): ex = {}
    for ft in FITTING_TYPES:
        ex.setdefault(ft, False)
    settings["exclude_fitting_types"] = ex
    return settings


def _normalized_settings_payload(settings: dict) -> dict:
    s = _deep_merge_defaults(settings, DEFAULT_SETTINGS)
    mats = s.get("material_types", [])
    if not isinstance(mats, list) or not mats:
        mats = MATERIAL_TYPE_PRESET
    s["material_types"] = sort_case_insensitive(
        dedupe_case_insensitive_keep_first([str(x).strip() for x in mats if str(x).strip()])
    )
    s["company_side_multipliers"] = _clean_multiplier_table(s.get("company_side_multipliers", {}))
    psm = s.get("project_side_multipliers", {})
    s["project_side_multipliers"] = {
        str(k).strip(): _clean_multiplier_table(v)
        for k, v in psm.items()
        if str(k).strip()
    }
    mode = str(s.get("multiplier_mode", "Company") or "Company")
    s["multiplier_mode"] = mode if mode in ("Company", "Project") else "Company"
    return ensure_company_defaults(s)


def _load_settings_from_file(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    return _normalized_settings_payload(payload)


def _write_json_file(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(f"{path.suffix}.tmp")
    tmp_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def load_settings() -> dict:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    primary = _load_settings_from_file(SETTINGS_FILE)
    if primary is not None:
        return primary
    backup = _load_settings_from_file(SETTINGS_BACKUP_FILE)
    if backup is not None:
        _write_json_file(SETTINGS_FILE, backup)
        return backup
    s = json.loads(json.dumps(DEFAULT_SETTINGS))
    return ensure_company_defaults(s)


def save_settings(settings: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    sanitized = ensure_company_defaults(dict(settings))
    sanitized.pop("admin_password", None)
    _write_json_file(SETTINGS_FILE, sanitized)
    _write_json_file(SETTINGS_BACKUP_FILE, sanitized)

# ============================
# Multiplier lookup
# ============================

def get_effective_multiplier(settings: dict, project: str, material: str, fitting: str) -> float:
    mat = str(material).strip()
    fit = str(fitting).strip()
    if project and project != "Company Default":
        psm = settings.get("project_side_multipliers", {})
        proj_tbl = _ci_get(psm, project)
        if isinstance(proj_tbl, dict):
            mp = _ci_get(proj_tbl, mat)
            if isinstance(mp, dict):
                v = _ci_get(mp, fit)
                if v is not None:
                    try: return round(float(v), 2)
                    except Exception: pass
    csm = settings.get("company_side_multipliers", {})
    if isinstance(csm, dict):
        mp = _ci_get(csm, mat)
        if isinstance(mp, dict):
            v = _ci_get(mp, fit)
            if v is not None:
                try: return round(float(v), 2)
                except Exception: pass
    return round(float(PROGRAM_DEFAULT_MULTIPLIER), 2)

# ============================
# Size parsing
# ============================

def _clean_size_text(s: str) -> str:
    s = clean_token(s).replace("ø","").replace("Ø","").replace('"',"").lower()
    s = re.sub(r"[^0-9x/\-.\s]"," ",s)
    return re.sub(r"\s+"," ",s).strip()


def parse_single_size(s: str) -> float:
    s = _clean_size_text(s)
    m = re.fullmatch(r"(\d+)\s+(\d+)\s*/\s*(\d+)", s)
    if m:
        w,n,d = m.groups(); return float(w)+float(n)/float(d)
    m = re.fullmatch(r"(\d+)\s*-\s*(\d+)\s*/\s*(\d+)", s)
    if m:
        w,n,d = m.groups(); return float(w)+float(n)/float(d)
    m = re.fullmatch(r"(\d+)\s*/\s*(\d+)", s)
    if m:
        n,d = m.groups(); return float(n)/float(d)
    m = re.fullmatch(r"\d+(\.\d+)?", s)
    if m:
        return float(s)
    raise ValueError(f"Unparseable size: {s}")


def size_to_diameter_in(size_str: str) -> float:
    s = _clean_size_text(size_str)
    s = re.sub(r"\s*x\s*"," x ",s)
    parts = [p.strip() for p in s.split(" x ") if p.strip()]
    if not parts:
        raise ValueError(f"No parsable size parts: {size_str}")
    parsed = []
    for p in parts:
        try:
            parsed.append(parse_single_size(p))
        except Exception:
            m = re.search(r"\d+(\.\d+)?|\d+\s*/\s*\d+|\d+\s*-\s*\d+\s*/\s*\d+|\d+\s+\d+\s*/\s*\d+", p)
            if m: parsed.append(parse_single_size(m.group(0)))
    if not parsed:
        raise ValueError(f"Could not parse any diameter from: {size_str}")
    return max(parsed)

# ============================
# Batch finding
# ============================

def _batch_from_filename(filename: str):
    m = BATCH_IN_FILENAME_RE.search(filename.upper())
    return m.group(0).upper() if m else None


def _find_batch_in_row(items: list, filename_batch):
    for tok in items:
        m = BATCH_IN_ROW_RE.search(tok.upper())
        if m: return m.group(0).upper()
    if filename_batch: return filename_batch
    for tok in reversed(items):
        if FALLBACK_BATCH_TOKEN_RE.fullmatch(tok.upper()): return tok.upper()
    return None

# ============================
# Row normalization
# ============================

def _split_length_and_trailing_material(token: str):
    raw = clean_token(token)
    m = LENGTH_RE.search(raw.replace("Ø", "").replace("ø", ""))
    if not m:
        return None, raw
    return m.group(0).strip(), raw[m.end():].strip()


def normalize_row_tokens(tokens: list) -> list:
    return [clean_token(t) for t in tokens if clean_token(t)]


def normalize_row(row: list, source_file: str, filename_batch, reject_log: list):
    items = [clean_token(c) for c in row if c is not None]
    items = [c for c in items if c and c.lower() != "none"]
    if not items:
        return None

    batch = _find_batch_in_row(items, filename_batch)
    if not batch:
        reject_log.append({"Reason": "Missing Batch", "Row Preview": " | ".join(items[:12])})
        return None

    idx_len = None
    length_token = None
    trailing_material_from_len = ""
    for i, tok in enumerate(items):
        tok2 = tok.replace("ø", "").replace("Ø", "")
        if LENGTH_RE.search(tok2):
            idx_len = i
            length_token, trailing_material_from_len = _split_length_and_trailing_material(tok2)
            if not length_token:
                length_token = tok2.strip()
                trailing_material_from_len = ""
            break

    if idx_len is None or not length_token:
        reject_log.append({"Reason": "Missing Length pattern", "Row Preview": " | ".join(items[:12])})
        return None

    count = None
    idx_count = None
    for j in range(idx_len - 1, -1, -1):
        if re.fullmatch(r"\d+", items[j]):
            count = int(items[j])
            idx_count = j
            break
    if idx_count is None:
        idx_count = idx_len

    after_len = items[idx_len + 1:]
    material_tokens = []
    if trailing_material_from_len:
        material_tokens.append(trailing_material_from_len)
    for tok in after_len:
        if BATCH_IN_ROW_RE.fullmatch(tok.upper()):
            continue
        material_tokens.append(tok)
    material = " ".join([t for t in material_tokens if t]).strip()
    if not material:
        reject_log.append({"Reason": "Missing Material segment", "Row Preview": " | ".join(items[:12])})
        return None

    left = items[:idx_count]
    if len(left) < 3:
        reject_log.append({"Reason": "Left side too short", "Row Preview": " | ".join(items[:12])})
        return None

    size = left[0].replace("ø", "").replace("Ø", "").replace('"', "").strip()
    install_type = left[1].strip()
    description = " ".join(left[2:]).strip()
    if not description:
        reject_log.append({"Reason": "Missing Description", "Row Preview": " | ".join(items[:12])})
        return None

    return {
        "Batch": batch,
        "Size": size,
        "Install Type": install_type,
        "Description": description,
        "Count": count,
        "Length": length_token,
        "Material": material,
        "Material Type": "",
        "Source File": source_file,
    }

# ============================
# PDF extraction
# ============================

def extract_from_pdf(pdf_path: Path, errors: list) -> pd.DataFrame:
    records = []
    pages = 0
    total_tables = 0
    total_table_rows_seen = 0
    filename_batch = _batch_from_filename(pdf_path.name)
    reject_samples = []
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            pages = len(pdf.pages)
            for p_idx, page in enumerate(pdf.pages):
                try:
                    tables = page.extract_tables()
                except Exception as exc:
                    errors.append({
                        "Stage": "PDF Extract",
                        "Source File": pdf_path.name,
                        "Issue": f"Page {p_idx + 1}: extract_tables() failed -> {type(exc).__name__}: {exc}",
                    })
                    continue
                total_tables += len(tables or [])
                for table in (tables or []):
                    total_table_rows_seen += len(table or [])
                    for row in (table or []):
                        if len(reject_samples) < 25:
                            rec = normalize_row(row, pdf_path.name, filename_batch, reject_samples)
                        else:
                            rec = normalize_row(row, pdf_path.name, filename_batch, [])
                        if rec:
                            records.append(rec)
    except Exception as exc:
        errors.append({
            "Stage": "PDF Open",
            "Source File": pdf_path.name,
            "Issue": f"pdfplumber.open() failed -> {type(exc).__name__}: {exc}",
        })
        return pd.DataFrame()

    if not records:
        if total_tables == 0:
            errors.append({
                "Stage": "PDF Extract",
                "Source File": pdf_path.name,
                "Issue": f"No tables detected. Pages={pages}. (Likely scanned/image PDF.)",
            })
        else:
            reason_counts = {}
            for rejected in reject_samples:
                reason = rejected["Reason"]
                reason_counts[reason] = reason_counts.get(reason, 0) + 1
            reason_summary = ", ".join(
                [f"{k}={v}" for k, v in sorted(reason_counts.items(), key=lambda x: -x[1])]
            )
            errors.append({
                "Stage": "PDF Extract",
                "Source File": pdf_path.name,
                "Issue": (
                    f"Tables detected but 0 rows matched. Pages={pages}, tables={total_tables}, "
                    f"rows_seen={total_table_rows_seen}. Top rejections: {reason_summary or 'None'}."
                ),
            })
            for rejected in reject_samples[:10]:
                errors.append({
                    "Stage": "Row Rejected",
                    "Source File": pdf_path.name,
                    "Issue": f"{rejected['Reason']} | Preview: {rejected['Row Preview']}",
                })
    return pd.DataFrame(records).drop_duplicates() if records else pd.DataFrame()


def records_from_pdf(pdf_path: str, settings: dict, legend_maps) -> list:
    errors = []
    df = extract_from_pdf(Path(pdf_path), errors)
    if df.empty:
        return []

    records = []
    for _, row in df.iterrows():
        material_type = material_type_from_material(row.get("Material", ""), settings.get("material_types", []))
        fitting_type = classify_fitting_type_with_legend(row.get("Description", ""), legend_maps)
        records.append({
            "batch": row.get("Batch", ""),
            "size": row.get("Size", ""),
            "install_type": row.get("Install Type", ""),
            "description": row.get("Description", ""),
            "count": row.get("Count"),
            "length": row.get("Length", ""),
            "material": row.get("Material", ""),
            "material_type": material_type,
            "fitting_type": fitting_type,
            "source_file": row.get("Source File", Path(pdf_path).name),
        })
    return records


def format_excel(path: Path) -> None:
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1 or ws.max_column < 1:
            continue
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, min(ws.max_row, 400) + 1):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 80)
        header_map = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

        def fmt_col(col_name, num_fmt):
            if col_name in header_map:
                c = header_map[col_name]
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, c).number_format = num_fmt

        fmt_col("Count", "0")
        fmt_col("Diameter (in)", "0.00")
        fmt_col("Side Multiplier", "0.00")
        fmt_col("Total Inches", "0.00")
        if ws.max_row >= 2:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table_name = f"Tbl_{re.sub(r'[^A-Za-z0-9]', '', sheet_name)[:20] or 'Sheet'}"
            existing = {t.displayName for t in ws._tables}
            if table_name in existing:
                table_name += "_1"
            tab = Table(displayName=table_name, ref=ref)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)
    wb.save(path)


def format_error_summary(errors: list, max_files: int = 10, max_lines_per_file: int = 10) -> str:
    if not errors:
        return "No errors."
    by_file = {}
    for error in errors:
        file_name = str(error.get("Source File") or "(unknown)")
        by_file.setdefault(file_name, []).append(error)

    def sev(items):
        score = 0
        for item in items:
            stage = str(item.get("Stage", "")).lower()
            if "pdf open" in stage:
                score += 100
            elif "pdf extract" in stage:
                score += 50
            elif "row rejected" in stage:
                score += 25
            else:
                score += 1
        return score

    sorted_files = sorted(by_file.items(), key=lambda kv: (sev(kv[1]), len(kv[1])), reverse=True)
    lines = ["ERROR DETAILS", "------"]
    for idx, (file_name, items) in enumerate(sorted_files):
        if idx >= max_files:
            lines.append(f"... plus {len(sorted_files) - max_files} more file(s).")
            break
        lines.append(f"\nFile: {file_name}  (issues: {len(items)})")
        seen = set()
        printed = 0
        for item in items:
            key = (str(item.get("Stage", "")), str(item.get("Issue", "")))
            if key in seen:
                continue
            seen.add(key)
            lines.append(f"  - [{key[0]}] {key[1]}")
            printed += 1
            if printed >= max_lines_per_file:
                if len(seen) < len(items):
                    lines.append("  ... more issues (see Errors sheet).")
                break
    return "\n".join(lines)

# ============================
# Main run function
# ============================

def run_bom(pdf_paths: list, settings: dict, export_filename: str, mode: str, project: str) -> dict:
    """
    Run the full BOM pipeline using the legacy desktop parser flow,
    adapted to the hosted web app's settings and export structure.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    settings = ensure_company_defaults(dict(settings))
    settings["multiplier_mode"] = mode if mode in ("Company", "Project") else "Company"
    settings["selected_project"] = project if settings["multiplier_mode"] == "Project" else ""

    errors = []
    dfs = []
    legend_maps = load_legend_maps()

    for pdf_path in pdf_paths:
        pdf_df = extract_from_pdf(Path(pdf_path), errors)
        if not pdf_df.empty:
            dfs.append(pdf_df)

    out_filename = str(export_filename or "BOM_Export").strip() or "BOM_Export"
    if not out_filename.endswith(".xlsx"):
        out_filename += ".xlsx"
    out_path = EXPORTS_DIR / out_filename

    if not dfs:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            pd.DataFrame(columns=[]).to_excel(writer, sheet_name="Master", index=False)
            pd.DataFrame(errors).to_excel(writer, sheet_name="Errors", index=False)
        format_excel(out_path)
        return {
            "ok": True,
            "total_inches": 0,
            "rows": 0,
            "ok_rows": 0,
            "warn_rows": 0,
            "err_rows": 0,
            "errors": len(errors),
            "output": str(out_path),
            "output_filename": out_filename,
            "summary": f"No rows extracted.\n\n{format_error_summary(errors)}",
        }

    df = pd.concat(dfs, ignore_index=True).drop_duplicates()
    df.insert(0, "Row ID", range(1, len(df) + 1))
    df["Row Status"] = "OK"
    df["Row Issues"] = ""

    df["Fitting Type"] = df["Description"].apply(
        lambda desc: classify_fitting_type_with_legend(desc, legend_maps)
    )

    ex_map = settings.get("exclude_fitting_types", {})
    df = df[~df["Fitting Type"].map(lambda x: bool(ex_map.get(x, False)))].copy()
    df["Row ID"] = range(1, len(df) + 1)

    df["Material Type"] = df["Material"].apply(
        lambda m: material_type_from_material(m, settings.get("material_types", MATERIAL_TYPE_PRESET))
    )

    diameters = []
    for idx, row in df.iterrows():
        try:
            diameter = size_to_diameter_in(row.get("Size", ""))
            diameters.append(round(float(diameter), 4))
        except Exception as exc:
            diameters.append(None)
            msg = f"Size parse failed -> {type(exc).__name__}: {exc}"
            df.at[idx, "Row Status"] = "Error"
            df.at[idx, "Row Issues"] = df.at[idx, "Row Issues"] + ((" | " if df.at[idx, "Row Issues"] else "") + msg)
            errors.append({
                "Stage": "Row Parse",
                "Source File": row.get("Source File", ""),
                "Row ID": int(df.at[idx, "Row ID"]),
                "Issue": msg,
            })
    df["Diameter (in)"] = diameters

    df["Side Multiplier"] = df.apply(
        lambda row: get_effective_multiplier(
            settings,
            project,
            str(row.get("Material Type", "UNKNOWN")),
            str(row.get("Fitting Type", "Unclassified")),
        ),
        axis=1,
    )

    total_inches = []
    for idx, row in df.iterrows():
        count = row.get("Count", None)
        diameter = row.get("Diameter (in)", None)
        if count is None or (isinstance(count, float) and pd.isna(count)):
            total_inches.append(None)
            msg = "Missing Count -> Total Inches not calculated"
            if df.at[idx, "Row Status"] != "Error":
                df.at[idx, "Row Status"] = "Warning"
            df.at[idx, "Row Issues"] = df.at[idx, "Row Issues"] + ((" | " if df.at[idx, "Row Issues"] else "") + msg)
            errors.append({
                "Stage": "Row Calc",
                "Source File": row.get("Source File", ""),
                "Row ID": int(df.at[idx, "Row ID"]),
                "Issue": msg,
            })
            continue
        if diameter is None or (isinstance(diameter, float) and pd.isna(diameter)):
            total_inches.append(None)
            msg = "Missing/Invalid Diameter -> Total Inches not calculated"
            df.at[idx, "Row Status"] = "Error"
            df.at[idx, "Row Issues"] = df.at[idx, "Row Issues"] + ((" | " if df.at[idx, "Row Issues"] else "") + msg)
            errors.append({
                "Stage": "Row Calc",
                "Source File": row.get("Source File", ""),
                "Row ID": int(df.at[idx, "Row ID"]),
                "Issue": msg,
            })
            continue
        try:
            inches = float(diameter) * float(count) * float(row.get("Side Multiplier", FALLBACK_MULTIPLIER))
            total_inches.append(round(inches, 2))
        except Exception as exc:
            total_inches.append(None)
            msg = f"Total Inches calc failed -> {type(exc).__name__}: {exc}"
            df.at[idx, "Row Status"] = "Error"
            df.at[idx, "Row Issues"] = df.at[idx, "Row Issues"] + ((" | " if df.at[idx, "Row Issues"] else "") + msg)
            errors.append({
                "Stage": "Row Calc",
                "Source File": row.get("Source File", ""),
                "Row ID": int(df.at[idx, "Row ID"]),
                "Issue": msg,
            })
    df["Total Inches"] = total_inches

    df = df.sort_values(
        ["Batch", "Material Type", "Fitting Type", "Size", "Install Type", "Description"]
    ).reset_index(drop=True)

    numeric_ti = pd.to_numeric(df["Total Inches"], errors="coerce").fillna(0)
    total_val = float(numeric_ti.sum())
    inches_by_mat = df.assign(_ti=numeric_ti).groupby("Material Type")["_ti"].sum().sort_values(ascending=False)

    total_row = {c: "" for c in df.columns}
    total_row["Batch"] = "TOTAL"
    total_row["Total Inches"] = round(total_val, 2)
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    errors_df = pd.DataFrame(errors)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Master", index=False)
        if not errors_df.empty:
            errors_df.to_excel(writer, sheet_name="Errors", index=False)
    format_excel(out_path)

    ok_rows = int((df["Row Status"] == "OK").sum()) if "Row Status" in df.columns else 0
    warn_rows = int((df["Row Status"] == "Warning").sum()) if "Row Status" in df.columns else 0
    err_rows = int((df["Row Status"] == "Error").sum()) if "Row Status" in df.columns else 0

    summary_lines = [
        f"{APP_NAME} - SUMMARY",
        f"Version: {APP_VERSION}",
        f"Multiplier Mode: {'One-Off (Project Override)' if mode == 'Project' else 'Company Wide (Default)'}",
        f"Project Override: {project if mode == 'Project' else '(none)'}",
        "------",
        f"PDFs processed: {len(pdf_paths)}",
        f"Rows exported (incl. issues): {len(df) - 1}",
        f"Row Status: OK={ok_rows}  Warning={warn_rows}  Error={err_rows}",
        f"Errors logged: {len(errors)}",
        "",
        f"TOTAL INCHES: {total_val:,.2f}",
        "",
        "TOTAL INCHES BY MATERIAL TYPE:",
    ]
    for material, value in inches_by_mat.items():
        summary_lines.append(f"  - {material}: {value:,.2f}")
    if errors:
        summary_lines += ["", format_error_summary(errors)]
    summary_lines += ["", "Output file:", str(out_path)]

    return {
        "ok": True,
        "total_inches": round(total_val, 2),
        "rows": len(df) - 1,
        "ok_rows": ok_rows,
        "warn_rows": warn_rows,
        "err_rows": err_rows,
        "errors": len(errors),
        "output": str(out_path),
        "output_filename": out_filename,
        "summary": "\n".join(summary_lines),
    }
